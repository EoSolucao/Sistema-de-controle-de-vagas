import flet as ft
import pandas as pd
import os
from datetime import datetime
import socket
from openpyxl import load_workbook

# Caminho do arquivo Excel
file_path = r"C:\Users\Eloizo\Desktop\Vagas\CadastroVagas.xlsx"


# Função para carregar os dados de apoio
def load_support_data():
    try:
        apoio_data = pd.read_excel(file_path, sheet_name="Apoio")
        return (
            apoio_data["Tipo de Contrato"].dropna().tolist(),
            apoio_data["Motivo"].dropna().tolist(),
        )
    except Exception as e:
        print(f"Erro ao carregar dados de apoio: {e}")
        return [], []


# Função para calcular dias em aberto
def calculate_days_open(data_abertura, data_fechamento):
    if pd.notnull(data_fechamento) and data_fechamento != "":
        return (
            pd.to_datetime(data_fechamento, format="%d/%m/%Y")
            - pd.to_datetime(data_abertura, format="%d/%m/%Y")
        ).days + 1
    return (datetime.now() - pd.to_datetime(data_abertura, format="%d/%m/%Y")).days + 1


# Função para salvar os dados na planilha
def save_data(
    vaga,
    quant_vagas,
    data_abertura,
    tipo_contrato,
    motivo,
    data_fechamento,
    status,
    id_vaga=None,
):
    try:
        # Carrega o workbook existente
        book = load_workbook(file_path)

        # Verifica se a planilha 'Cadastro' existe
        if "Cadastro" in book.sheetnames:
            # Se existir, carrega os dados
            df = pd.read_excel(file_path, sheet_name="Cadastro")
        else:
            # Se não existir, cria um novo DataFrame
            df = pd.DataFrame(
                columns=[
                    "Id Vaga",
                    "Vaga",
                    "Quant. de Vagas",
                    "Data Abertura",
                    "Tipo de Contrato",
                    "Motivo",
                    "Data Fechamento",
                    "Status",
                    "Dias em Aberto",
                    "Data Alteração",
                    "Nome Máquina",
                ]
            )

        if id_vaga:
            # Atualiza a linha existente
            row_index = df.index[df["Id Vaga"] == id_vaga].tolist()[0]
            df.loc[row_index, "Vaga"] = vaga
            df.loc[row_index, "Quant. de Vagas"] = quant_vagas
            df.loc[row_index, "Data Abertura"] = data_abertura
            df.loc[row_index, "Tipo de Contrato"] = tipo_contrato
            df.loc[row_index, "Motivo"] = motivo
            df.loc[row_index, "Data Fechamento"] = data_fechamento
            df.loc[row_index, "Status"] = "Aberto" if not data_fechamento else "Fechado"
            df.loc[row_index, "Dias em Aberto"] = calculate_days_open(
                data_abertura, data_fechamento
            )
            df.loc[row_index, "Data Alteração"] = datetime.now().strftime(
                "%d/%m/%Y %H:%M:%S"
            )
            df.loc[row_index, "Nome Máquina"] = socket.gethostname()
        else:
            # Adiciona uma nova linha
            new_id = df["Id Vaga"].max() + 1 if not df.empty else 1
            new_row = {
                "Id Vaga": new_id,
                "Vaga": vaga,
                "Quant. de Vagas": quant_vagas,
                "Data Abertura": data_abertura,
                "Tipo de Contrato": tipo_contrato,
                "Motivo": motivo,
                "Data Fechamento": data_fechamento,
                "Status": "Aberto" if not data_fechamento else "Fechado",
                "Dias em Aberto": calculate_days_open(data_abertura, data_fechamento),
                "Data Alteração": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "Nome Máquina": socket.gethostname(),
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        # Salva o DataFrame atualizado na planilha 'Cadastro'
        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name="Cadastro", index=False)

        return True
    except Exception as e:
        print(f"Erro ao salvar dados: {e}")
        return False


# Função para carregar os dados da planilha
def load_data():
    try:
        df = pd.read_excel(file_path, sheet_name="Cadastro")
        return df
    except Exception as e:
        print(f"Erro ao carregar dados: {e}")
        return pd.DataFrame()


# Função para filtrar dados
def filter_data(search_value):
    df = load_data()
    if search_value.isdigit():
        filtered_df = df[df["Id Vaga"] == int(search_value)]
    else:
        filtered_df = df[df["Vaga"].str.contains(search_value, case=False, na=False)]
    return filtered_df


def cadastro_vagas_page(page: ft.Page):
    # Carrega os dados de apoio
    tipo_contrato_options, motivo_options = load_support_data()

    # Cria os campos de entrada
    vaga_input = ft.TextField(label="Vaga", width=300)
    quant_vagas_input = ft.TextField(label="Quant. de Vagas", width=100)
    data_abertura_input = ft.TextField(label="Data Abertura (dd/mm/aaaa)", width=200)
    tipo_contrato_select = ft.Dropdown(
        label="Tipo de Contrato",
        options=[ft.dropdown.Option(tc) for tc in tipo_contrato_options],
        width=200,
    )
    motivo_select = ft.Dropdown(
        label="Motivo",
        options=[ft.dropdown.Option(m) for m in motivo_options],
        width=200,
    )
    data_fechamento_input = ft.TextField(
        label="Data Fechamento (dd/mm/aaaa)", width=200
    )
    search_input = ft.TextField(label="Buscar por ID ou Vaga", width=200)

    current_id = ft.Text()

    # Função para limpar os campos de entrada
    def clear_inputs():
        vaga_input.value = ""
        quant_vagas_input.value = ""
        data_abertura_input.value = ""
        tipo_contrato_select.value = None
        motivo_select.value = None
        data_fechamento_input.value = ""
        current_id.value = ""
        page.update()

    # Função chamada ao clicar no botão salvar
    def save_clicked(e):
        if current_id.value:
            show_confirmation_dialog("Deseja fazer esta alteração?", confirm_save)
        else:
            confirm_save(True)

    # Função para confirmar o salvamento
    def confirm_save(confirmed):
        if confirmed:
            success = save_data(
                vaga_input.value,
                quant_vagas_input.value,
                data_abertura_input.value,
                tipo_contrato_select.value,
                motivo_select.value,
                data_fechamento_input.value,
                "",
                int(current_id.value) if current_id.value else None,
            )
            if success:
                show_success_dialog("Salvo com sucesso!")
                clear_inputs()
                update_table()
        if page.dialog:
            page.dialog.open = False
        page.update()

    # Função para mostrar diálogo de confirmação
    def show_confirmation_dialog(message, on_result):
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirmação"),
            content=ft.Text(message),
            actions=[
                ft.TextButton("Sim", on_click=lambda _: on_result(True)),
                ft.TextButton("Não", on_click=lambda _: on_result(False)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.dialog = dialog
        dialog.open = True
        page.update()

    # Função para mostrar diálogo de sucesso
    def show_success_dialog(message):
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Sucesso"),
            content=ft.Text(message),
            actions=[
                ft.TextButton("OK", on_click=lambda _: close_dialog()),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.dialog = dialog
        dialog.open = True
        page.update()

    # Função para fechar o diálogo
    def close_dialog():
        if page.dialog:
            page.dialog.open = False
        page.update()

    save_button = ft.ElevatedButton(text="Salvar", on_click=save_clicked)

    # Cria a tabela de dados
    data_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("Id Vaga")),
            ft.DataColumn(ft.Text("Vaga")),
            ft.DataColumn(ft.Text("Quant. de Vagas")),
            ft.DataColumn(ft.Text("Data Abertura")),
            ft.DataColumn(ft.Text("Tipo de Contrato")),
            ft.DataColumn(ft.Text("Motivo")),
            ft.DataColumn(ft.Text("Data Fechamento")),
            ft.DataColumn(ft.Text("Status")),
            ft.DataColumn(ft.Text("Dias em Aberto")),
            ft.DataColumn(ft.Text("Editar")),
        ],
        rows=[],
    )

    # Função para editar uma linha
    def edit_row(e):
        row = e.control.data
        fill_form(row)

    # Função para atualizar a tabela
    def update_table(e=None):
        df = filter_data(search_input.value) if search_input.value else load_data()
        data_table.rows.clear()
        for _, row in df.iterrows():
            dias_em_aberto = row["Dias em Aberto"]
            status_color = "blue" if row["Status"] == "Aberto" else "green"
            dias_color = (
                "red"
                if dias_em_aberto > 20
                else (
                    "blue"
                    if dias_em_aberto > 15
                    else "green" if dias_em_aberto < 14 else "black"
                )
            )

            edit_button = ft.IconButton(
                icon=ft.icons.EDIT,
                tooltip="Editar",
                on_click=edit_row,
                data=row,
            )

            new_row = ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(row["Id Vaga"])),
                    ft.DataCell(ft.Text(row["Vaga"])),
                    ft.DataCell(ft.Text(row["Quant. de Vagas"])),
                    ft.DataCell(
                        ft.Text(
                            row["Data Abertura"].strftime("%d/%m/%Y")
                            if isinstance(row["Data Abertura"], datetime)
                            else row["Data Abertura"]
                        )
                    ),
                    ft.DataCell(ft.Text(row["Tipo de Contrato"])),
                    ft.DataCell(ft.Text(row["Motivo"])),
                    ft.DataCell(
                        ft.Text(
                            row["Data Fechamento"].strftime("%d/%m/%Y")
                            if isinstance(row["Data Fechamento"], datetime)
                            and pd.notnull(row["Data Fechamento"])
                            else ""
                        )
                    ),
                    ft.DataCell(ft.Text(row["Status"], color=status_color)),
                    ft.DataCell(ft.Text(f"{dias_em_aberto:.0f}", color=dias_color)),
                    ft.DataCell(edit_button),
                ]
            )
            data_table.rows.append(new_row)
        page.update()

    # Conecta a função de atualização ao evento on_change do campo de busca
    search_input.on_change = update_table

    # Função para preencher o formulário com os dados de uma linha
    def fill_form(row):
        current_id.value = str(row["Id Vaga"])
        vaga_input.value = row["Vaga"]
        quant_vagas_input.value = str(row["Quant. de Vagas"])
        data_abertura_input.value = (
            row["Data Abertura"].strftime("%d/%m/%Y")
            if isinstance(row["Data Abertura"], datetime)
            else row["Data Abertura"]
        )
        tipo_contrato_select.value = row["Tipo de Contrato"]
        motivo_select.value = row["Motivo"]
        data_fechamento_input.value = (
            row["Data Fechamento"].strftime("%d/%m/%Y")
            if isinstance(row["Data Fechamento"], datetime)
            and pd.notnull(row["Data Fechamento"])
            else ""
        )
        page.update()

    # Cria um container scrollable para a tabela
    scrollable_table = ft.Container(
        content=ft.Column(
            controls=[data_table],
            scroll=ft.ScrollMode.ALWAYS,
        ),
        width=1600,
        height=1000,
    )

    # Cria o layout da página
    layout = ft.Column(
        [
            ft.Text(
                "Cadastro de Vagas",
                size=24,
                weight="bold",
                text_align=ft.TextAlign.CENTER,
            ),
            ft.Row(
                controls=[
                    vaga_input,
                    quant_vagas_input,
                    data_abertura_input,
                    tipo_contrato_select,
                    motivo_select,
                    data_fechamento_input,
                    search_input,
                ]
            ),
            save_button,
            scrollable_table,
        ]
    )

    # Atualiza a tabela inicialmente
    update_table()

    # Retorna o layout da página de cadastro de vagas
    return layout


def processo_seletivo_page(page: ft.Page):
    return ft.Column(
        [
            ft.Text("Processo Seletivo", size=24, weight="bold"),
            ft.Text("Esta página está em desenvolvimento."),
        ]
    )


def main(page: ft.Page):
    page.title = "Sistema de Controle de Vagas"
    page.window_width = 1600
    page.window_height = 1100

    def route_change(route):
        page.views.clear()
        tabs = ft.Tabs(
            selected_index=0,
            tabs=[
                ft.Tab(text="Cadastro de Vagas", content=cadastro_vagas_page(page)),
                ft.Tab(text="Processo Seletivo", content=processo_seletivo_page(page)),
            ],
        )
        page.views.append(
            ft.View(
                "/",
                [
                    ft.AppBar(
                        title=ft.Text("Sistema de Controle de Vagas"),
                        bgcolor=ft.colors.SURFACE_VARIANT,
                    ),
                    tabs,
                ],
            )
        )
        page.update()

    def view_pop(view):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)


ft.app(target=main)
